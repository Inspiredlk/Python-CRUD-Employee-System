from flask import Flask, request, redirect, url_for, render_template_string, send_file
from flask_sqlalchemy import SQLAlchemy
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from datetime import datetime
import pandas as pd 
from io import BytesIO



app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///users.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)

# -----------------------------
# Database Model
# -----------------------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(30), nullable=True)


with app.app_context():
    db.create_all()

# -----------------------------
# Bootstrap Templates (Bigger text + colored buttons + Icons)
# -----------------------------
TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Employee Management</title>

<!-- Bootstrap CSS -->
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">

<!-- Animate CSS -->
<link href="https://cdnjs.cloudflare.com/ajax/libs/animate.css/4.1.1/animate.min.css" rel="stylesheet">

<!-- Bootstrap Icons -->
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.5/font/bootstrap-icons.css">

<style>
@keyframes gradientBG {
  0% {background-position: 0% 50%;}
  25% {background-position: 50% 100%;}
  50% {background-position: 100% 50%;}
  75% {background-position: 50% 0%;}
  100% {background-position: 0% 50%;}
}
#split-text span {
  display: inline-block;
  font-size: 34px;
  font-weight: bold;
  opacity: 0;
  transform: translateY(-50px) rotate(-30deg);
  animation: flyIn 1s forwards;
}

/* Add delay for each letter */
#split-text span:nth-child(1) { animation-delay: 0s; }
#split-text span:nth-child(2) { animation-delay: 0.05s; }
#split-text span:nth-child(3) { animation-delay: 0.1s; }

@keyframes flyIn {
  to {
    opacity: 1;
    transform: translateY(0) rotate(0deg);
  }
}

body {
  background: linear-gradient(-45deg, #FFA500, #000000, #000000, #FFA500);
  background-size: 400% 400%;
  animation: gradientBG 20s ease infinite; /* faster and smoother */
  
}


.navbar {
  backdrop-filter: blur(10px);
  background-color: #1a1a1a;
  border-radius: 15px;     
  padding: 30px 0;
}

.navbar-brand {
  font-size: 34px;
  letter-spacing: 1px;
  justify-content: center;
  align-items: center;
}

.card {
  background: #001B3A;           /* dark blue background */
  backdrop-filter: blur(15px);   /* glass effect */
  border-radius: 15px;           /* rounded corners */
  border: 1px solid #ffffff;     /* solid white border */
  transition: transform 0.3s ease, box-shadow 0.3s ease;
  color: #ffffff;                /* text color */
  padding: 20px;                 /* optional padding inside card */
}

.card h4 { font-size: 26px; }
.card input.form-control {
  font-size: 16px;
  background-color: rgba(255,255,255,0.1);
  color: #fff;
  border: 1px solid rgba(255,255,255,0.3);
}
.card input::placeholder { color: rgba(255,255,255,0.7); }

.card button.btn {
  font-size: 16px;
  background-color: #FFA500;
  color: #000000;
  border: none;
  transition: transform 0.3s ease, box-shadow 0.3s ease;
  font-weight: bold;
}
.card button.btn:hover {
  transform: scale(1.05);
  box-shadow: 0 0 15px #FF7F00;
}

.table {
  --bs-table-bg: transparent !important;
  background-color: transparent !important;
  color: white !important;
}
.table > :not(caption) > * > * {
  background-color: transparent !important;
  color: white !important;
  border-color: rgba(255,255,255,0.1) !important;
}
.table thead { background: #FFA500 }
.table thead th { color: black !important; font-weight: bold; }
.table tbody tr:hover { background: rgba(255,255,255,0.05); }

input { background: rgba(255,255,255,0.1) !important; color: white !important; border: 1px solid rgba(255,255,255,0.3) !important; }
input::placeholder { color: #ccc !important; }

.alert-danger { background: rgba(255,0,0,0.2); border: none; }
</style>
</head>

<body>

<nav class="navbar navbar-expand-lg mb-4">
  <div class="container-fluid px-5 d-flex justify-content-center">
    <span class="navbar-brand text-white" id="split-text">
      <span>🚀</span>
      <span>E</span><span>m</span><span>p</span><span>l</span><span>o</span><span>y</span><span>e</span><span>e</span>
      <span>M</span><span>a</span><span>n</span><span>a</span><span>g</span><span>e</span><span>m</span><span>e</span><span>n</span><span>t</span>
      <span>S</span><span>y</span><span>s</span><span>t</span><span>e</span><span>m</span>
    </span>
  </div>
</nav>

<div class="container animate__animated animate__fadeIn">

{% if error %}
<div class="alert alert-danger text-center">{{ error }}</div>
{% endif %}

<div class="row g-4">

<div class="col-lg-4">
<div class="card p-4 animate__animated animate__fadeInLeft">
<h4 class="mb-3">Add Employee</h4>


<form method="post" action="{{ url_for('create_user') }}">
<div class="mb-3">
<input class="form-control" name="full_name" required placeholder="Full Name">
</div>

<div class="mb-3">
<input class="form-control" name="email" type="email" required placeholder="Email">
</div>

<div class="mb-3">
<input class="form-control" name="phone" placeholder="Phone">
</div>

<button class="btn btn-primary w-100">Add User</button>
</form>

</div>
</div>

<div class="col-lg-8">
<div class="card p-4 animate__animated animate__fadeInRight">
<div class="d-flex justify-content-between mb-3">
<h4>Employees</h4>
<span class="badge" style="color: #28a745;font-size: 24px;font-weight: bold;">Total: {{ "%02d"|format(users|length) }}</span>

</div>

<div class="table-responsive">
<table class="table align-middle">
<thead>
<tr>
<th>ID</th>
<th>Name</th>
<th>Email</th>
<th>Phone</th>
<th>Actions</th>
</tr>
</thead>

<tbody>
{% for u in users %}
<tr>
<td>{{u.id}}</td>
<td>{{u.full_name}}</td>
<td>{{u.email}}</td>
<td>{{u.phone or ""}}</td>

<td class="d-flex gap-2">
<!-- Edit Icon -->
<a class="btn btn-outline-light btn-sm" href="{{ url_for('edit_form', user_id=u.id) }}" title="Edit">
    <i class="bi bi-pencil-square"></i>
</a>

<!-- Delete Icon -->
<a class="btn btn-outline-danger btn-sm" href="{{ url_for('delete_user', user_id=u.id) }}" onclick="return confirm('Delete user?')" title="Delete">
    <i class="bi bi-trash"></i>
</a>
</td>

</tr>
{% endfor %}

{% if users|length == 0 %}
<tr>
<td colspan="5" class="text-center text-muted">No employees yet.</td>
</tr>
{% endif %}
</tbody>
</table>
</div>

</div>
</div>

</div>

<!-- Button to export records to Excel -->
<div class="d-flex justify-content-end mt-4"> 
    <a href="{{ url_for('view_records') }}" 
       class="btn" 
       style="background-color: #FFA500; color: #000000; font-weight: bold;">
      View All Records
    </a>
</div>


</div>
</div>

</body>
</html>
"""

# -----------------------------
# Routes
# -----------------------------
@app.get("/")
def index():
    users = User.query.order_by(User.id.desc()).all()
    error = request.args.get("error")
    return render_template_string(TEMPLATE, users=users, error=error)


@app.post("/create")
def create_user():
    full_name = request.form["full_name"].strip()
    email = request.form["email"].strip().lower()
    phone = request.form.get("phone", "").strip()

    if User.query.filter_by(email=email).first():
        return redirect(url_for("index", error="Email already exists. Please use another email."))

    db.session.add(User(full_name=full_name, email=email, phone=phone))
    db.session.commit()
    return redirect(url_for("index"))

@app.get("/view")
def view_records():
    users = User.query.all()
    
    records = [
        {"ID": u.id, "Full Name": u.full_name, "Email": u.email, "Phone": u.phone or "N/A"} 
        for u in users
    ]
    
    EXCEL_HTML_TEMPLATE = """
    <!doctype html>
    <html lang="en">
    <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Employee Records</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: #0a0a0a; /* dark background */
            color: #ffffff;       /* white text */
            padding: 30px;
        }
        h2 {
            text-align: center;
            margin-bottom: 25px;
            color: #FFA500; /* orange heading */
        }
.table-container {
    background: #001B3A;  /* dark card background */
    padding: 20px;
    border-radius: 10px;
    overflow-x: auto;
    box-shadow: 0px 0px 15px rgba(0,0,0,0.5);
    width: 70%;           /* reduced width */
    margin: 0 auto;       /* center container */
}

table {
    border-collapse: collapse;
    width: 100%;           /* table fills container */
}

table th, table td {
    border: 1px solid #444;
    padding: 10px 12px;
    text-align: left;
    font-size: 14px;
    color: #ffffff;
}

table thead {
    background-color: #FFA500;
    color: #000;
    font-weight: bold;
}

table tbody tr:nth-child(even) {
    background-color: #204060;
}

table tbody tr:hover {
    background-color: #001B3A;


        }
        .btn-download {
            background-color: #28a745;
            color: #fff;
            font-weight: bold;
            border: none;
            margin-top: 20px;
            padding: 10px 15px;
            width: auto;          /* smaller width */
            display: inline-block; /* not full width */
        }
        .btn-download:hover {
            background-color: #218838;
        }
    </style>
    </head>
    <body>
        <h2>All Employee Records</h2>
        <div class="table-container">
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Full Name</th>
                        <th>Email</th>
                        <th>Phone</th>
                    </tr>
                </thead>
                <tbody>
                {% for row in records %}
                    <tr>
                        <td>{{ row['ID'] }}</td>
                        <td>{{ row['Full Name'] }}</td>
                        <td>{{ row['Email'] }}</td>
                        <td>{{ row['Phone'] }}</td>
                    </tr>
                {% endfor %}
                {% if records|length == 0 %}
                    <tr>
                        <td colspan="4" class="text-center">No employees yet.</td>
                    </tr>
                {% endif %}
                </tbody>
            </table>
        </div>
        <div class="d-flex justify-content-center gap-3 mt-4">
            <a href="{{ url_for('export_to_excel') }}"
   class="btn download-btn"
   style="background-color:#28a745; color:white;">
   Download Excel
</a>

<a href="{{ url_for('export_to_pdf') }}"
   class="btn download-btn"
   style="background-color:#dc3545; color:white;">
   Download PDF
</a>

</div>
             
        </div>
    </body>
    </html>
    """
    
    return render_template_string(EXCEL_HTML_TEMPLATE, records=records)

# Export to Excel
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

@app.route("/export")
def export_to_excel():
    users = User.query.all()
    data = {
        "ID": [user.id for user in users],
        "Full Name": [user.full_name for user in users],
        "Email": [user.email for user in users],
        "Phone": [user.phone or "N/A" for user in users],
    }
    df = pd.DataFrame(data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Employees")
        ws = writer.sheets["Employees"]

        # Styling headers
        header_font = Font(bold=True, color="FFFFFFFF")  # white bold
        header_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")  # orange
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Styling all cells with borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

    output.seek(0)
    return send_file(
        output, 
        as_attachment=True, 
        download_name="employees.xlsx", 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from io import BytesIO

@app.get("/export_pdf")
def export_to_pdf():

    users = User.query.all()

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)

    width, height = letter

    # ---------------- HEADER ----------------
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 50,
                          "Employee Report")

    pdf.setFont("Helvetica", 10)
    pdf.drawRightString(width - 40, height - 70,
                        datetime.now().strftime("%Y-%m-%d"))

    y = height - 120

    # ---------------- EMPLOYEE DETAILS ----------------
    pdf.setFont("Helvetica", 12)

    for user in users:

        pdf.drawString(80, y, f"Name : {user.full_name}")
        y -= 20

        pdf.drawString(80, y, f"Email : {user.email}")
        y -= 20

        pdf.drawString(80, y, f"Phone Number : {user.phone or 'N/A'}")
        y -= 30

        # line separator
        pdf.line(60, y, width - 60, y)
        y -= 30

        # New page if space finished
        if y < 100:
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y = height - 100

    # ---------------- FOOTER ----------------
    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawCentredString(
        width / 2,
        30,
        "Generated by Employee Management System"
    )

    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="employees_report.pdf",
        mimetype="application/pdf"
    )

# Placeholder EDIT_TEMPLATE (you can design similar to main template)
EDIT_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Edit Employee</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.5/font/bootstrap-icons.css">
<style>
body { background: #0f2027; color: white; font-family: 'Segoe UI', sans-serif; }
.card { background: #001B3A; padding: 30px; border-radius: 15px; }
</style>
</head>
<body>
<div class="container mt-5">
<div class="card mx-auto" style="max-width: 500px;">
<h4 class="mb-3">Edit Employee</h4>
<form method="post" action="{{ url_for('edit_user', user_id=u.id) }}">
<div class="mb-3">
<input class="form-control" name="full_name" value="{{u.full_name}}" required>
</div>
<div class="mb-3">
<input class="form-control" name="email" type="email" value="{{u.email}}" required>
</div>
<div class="mb-3">
<input class="form-control" name="phone" value="{{u.phone}}">
</div>
<button class="btn btn-warning w-100">Update User</button>
</form>
</div>
</div>
</body>
</html>

"""
# -----------------------------
# Excel Page Template
# -----------------------------
EXCEL_TEMPLATE = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>All Employee Records</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
body {
    background: linear-gradient(-45deg, #FFA500, #000000, #000000, #FFA500);
    background-size: 400% 400%;
    animation: gradientBG 20s ease infinite;
    font-family: 'Segoe UI', sans-serif;
    color: #000;
}
@keyframes gradientBG {
  0% {background-position: 0% 50%;}
  25% {background-position: 50% 100%;}
  50% {background-position: 100% 50%;}
  75% {background-position: 50% 0%;}
  100% {background-position: 0% 50%;}
}
.table-container {
    background: #fff;
    padding: 20px;
    border-radius: 10px;
    overflow-x: auto;
}
table {
    border-collapse: collapse !important;
    width: 100%;
}
table th, table td {
    border: 1px solid #000;
    padding: 8px 12px;
    text-align: left;
}
table thead {
    background-color: #FFA500;
    color: #000;
    font-weight: bold;
}
table tbody tr:nth-child(even) {
    background-color: #f2f2f2;
}
table tbody tr:hover {
    background-color: #ddd;
}
.btn-download {
    background-color: #28a745;
    color: #fff;
    font-weight: bold;
    border: none;
    margin-top: 15px;
}
.btn-download:hover {
    background-color: #218838;
}
</style>
</head>
<body>
<div class="container mt-4">
<h2 class="mb-4 text-center">All Employee Records</h2>

<div class="table-container">
<table class="table table-bordered">
<thead>
<tr>
<th>ID</th>
<th>Full Name</th>
<th>Email</th>
<th>Phone</th>
</tr>
</thead>
<tbody>
{% for row in records %}
<tr>
<td>{{row['ID']}}</td>
<td>{{row['Full Name']}}</td>
<td>{{row['Email']}}</td>
<td>{{row['Phone']}}</td>
</tr>
{% endfor %}
{% if records|length == 0 %}
<tr>
<td colspan="4" class="text-center">No employees yet.</td>
</tr>
{% endif %}
</tbody>
</table>
</div>

<a href="{{ url_for('download_excel') }}" class="btn btn-download w-100">Download Excel</a>
<a href="{{ url_for('export_to_pdf') }}"
   class="btn"
   style="
       background-color:#dc3545;
       color:white;
       font-weight:bold;
       width:180px;
       margin-left:20px;
       padding:10px;
   ">
   Download PDF
</a>
</div>
</body>
</html>
"""



@app.get("/edit/<int:user_id>")
def edit_form(user_id):
    u = User.query.get_or_404(user_id)
    return render_template_string(EDIT_TEMPLATE, u=u)


@app.post("/edit/<int:user_id>")
def edit_user(user_id):
    u = User.query.get_or_404(user_id)
    new_email = request.form["email"].strip().lower()
    existing = User.query.filter_by(email=new_email).first()
    if existing and existing.id != user_id:
        return redirect(url_for("index", error="That email is already used by another user."))
    u.full_name = request.form["full_name"].strip()
    u.email = new_email
    u.phone = request.form.get("phone", "").strip()
    db.session.commit()
    return redirect(url_for("index"))


@app.get("/delete/<int:user_id>")
def delete_user(user_id):
    u = User.query.get_or_404(user_id)
    db.session.delete(u)
    db.session.commit()
    return redirect(url_for("index"))


# -----------------------------
# Run App
# -----------------------------
if __name__ == "__main__":
    app.run(debug=True)

    print("CI Testing")

    # CI Demostration 1